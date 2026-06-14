"""
P&ID Parser Demo - KU Automation (v2.0)
Extracts equipment, valves, and line data from P&ID diagrams using AI

IMPROVEMENTS in v2.0:
- PDF support via pdf2image conversion
- Multi-page PDF processing
- Tiled analysis for large/detailed images
- Improved prompts for precision extraction
- Result deduplication and merging
- Better error handling
"""

import os

import json
import base64
import tempfile
import io
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from collections import defaultdict

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Image processing
from PIL import Image

# PDF processing - optional, graceful fallback
try:
    from pdf2image import convert_from_bytes
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("⚠️  pdf2image not installed. PDF support disabled. Install with: pip install pdf2image")
    print("   Also requires poppler: brew install poppler (macOS) or apt install poppler-utils (Linux)")

app = FastAPI(title="P&ID Parser - KU Automation v2.0")

# CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# OpenAI client - initialized lazily
_client = None

def get_openai_client():
    global _client
    if _client is None:
        # Try multiple env var names (Railway shared variable workaround)
        api_key = os.environ.get("MYKEY_XYZ123") or os.environ.get("KU_OPENAI_SECRET") or os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise HTTPException(500, "OPENAI_API_KEY not configured")
        _client = OpenAI(api_key=api_key)
    return _client


# ============================================================================
# IMPROVED EXTRACTION PROMPTS
# ============================================================================

PID_EXTRACTION_PROMPT = """You are an expert P&ID (Piping and Instrumentation Diagram) analyzer with deep knowledge of oil & gas, chemical, and process engineering standards (ISA, ISO, ANSI).

TASK: Analyze this P&ID image section and extract ALL visible components with PRECISION.

## CRITICAL INSTRUCTIONS:
1. ONLY extract what you can CLEARLY see - do not infer or guess
2. Read ALL text labels, tags, and annotations carefully
3. For partial/unclear text, use "[unclear]" suffix
4. Equipment tags follow patterns: V-XXX (vessels), P-XXX (pumps), E-XXX (exchangers), C-XXX (compressors), T-XXX (tanks)
5. Instrument tags follow ISA standards: First letter = measured variable (P=pressure, T=temp, F=flow, L=level), subsequent = function
6. Line numbers typically: SIZE-SERVICE-NUMBER (e.g., 4"-P-101, 6"-CW-201)

## EXTRACT THESE CATEGORIES:

### 1. EQUIPMENT (vessels, pumps, exchangers, compressors, tanks, reactors, columns, drums, filters)
For each:
- tag: Exact tag as shown (e.g., "V-101", "P-102A/B")
- type: Equipment type (Vessel, Pump, Heat Exchanger, etc.)
- description: Name/service if labeled
- size: Dimensions or capacity if shown
- material: If specified (CS, SS, etc.)

### 2. VALVES (all valve types visible)
For each:
- tag: Exact tag (XV, PV, HV, CV, PSV, etc.)
- type: Gate, Globe, Ball, Check, Control, Relief, Butterfly, Plug, Needle
- size: Pipe size if shown
- line_number: Which line it's on
- fail_position: FC/FO/FL if shown for control valves

### 3. INSTRUMENTS (all instrumentation)
For each:
- tag: Full ISA tag (e.g., "PT-101", "FIC-102", "LSHH-201")
- type: Measured variable (Pressure, Temperature, Flow, Level, Analytical)
- function: Transmitter, Indicator, Controller, Switch, Alarm, Element
- location: Field/Panel if indicated

### 4. LINES/PIPING
For each distinct line:
- line_number: Full designation
- size: Nominal diameter
- service: Fluid/service code
- spec: Piping spec/class if shown
- insulation: Type/thickness if indicated
- from_equipment: Source
- to_equipment: Destination

### 5. NOTES & SPECIFICATIONS
- Design conditions (pressure, temperature)
- Material specs
- Hazardous area classification
- Any other callouts

## OUTPUT FORMAT:
Return ONLY valid JSON:
{
  "equipment": [...],
  "valves": [...],
  "instruments": [...],
  "lines": [...],
  "notes": [...],
  "summary": "Brief description of this P&ID section"
}

Be thorough but conservative - quality over quantity. Missing an item is better than inventing one."""


PID_FOCUSED_EQUIPMENT_PROMPT = """Analyze this P&ID image and extract ONLY EQUIPMENT (major process equipment).

Look for: Vessels, Pumps, Heat Exchangers, Compressors, Tanks, Reactors, Columns, Drums, Filters, Blowers, Turbines, Ejectors

For EACH piece of equipment, extract:
- tag: Exact tag number (e.g., "V-101", "P-102A/B", "E-201")
- type: Equipment type
- description: Name or service description
- size: Dimensions, capacity, or rating
- material: Construction material if shown

Return JSON: {"equipment": [...], "count": N}
ONLY output the JSON, nothing else."""


PID_FOCUSED_VALVES_PROMPT = """Analyze this P&ID image and extract ONLY VALVES.

Look for all valve symbols: Gate, Globe, Ball, Check, Control (with actuators), Relief/Safety (PSV), Butterfly, Plug, Needle, 3-way, Isolation

For EACH valve, extract:
- tag: Exact tag (e.g., "XV-101", "PV-102", "PSV-103", "HV-104")
- type: Valve type
- size: If shown
- line_number: Which line it's installed on
- actuator: Manual, Pneumatic, Electric, Hydraulic
- fail_position: FC/FO/FL for control valves

Return JSON: {"valves": [...], "count": N}
ONLY output the JSON, nothing else."""


PID_FOCUSED_INSTRUMENTS_PROMPT = """Analyze this P&ID image and extract ONLY INSTRUMENTS.

Look for: Transmitters, Indicators, Controllers, Switches, Alarms, Elements, Analyzers
ISA tag patterns: PT (pressure transmitter), TT (temp), FT (flow), LT (level), AT (analyzer), PI (pressure indicator), etc.

For EACH instrument, extract:
- tag: Full ISA tag (e.g., "PT-101", "FIC-102", "LSHH-201", "TE-301")
- type: Measured variable (Pressure, Temperature, Flow, Level, Analytical)
- function: Transmitter, Indicator, Controller, Switch, Alarm, Element
- location: Field or Panel/DCS if indicated
- setpoint: If shown

Return JSON: {"instruments": [...], "count": N}
ONLY output the JSON, nothing else."""


PID_FOCUSED_LINES_PROMPT = """Analyze this P&ID image and extract ONLY PIPING/LINES.

Look for: Main process lines, utility lines, instrument air, nitrogen, steam, cooling water
Line number format typically: SIZE"-SERVICE-NUMBER (e.g., 4"-P-101, 6"-CW-201)

For EACH distinct line, extract:
- line_number: Full line designation
- size: Nominal pipe size
- service: Fluid/service (Process, CW, Steam, IA, N2, etc.)
- spec: Piping class/spec if shown
- insulation: Type and thickness if shown
- from_equipment: Where line originates
- to_equipment: Where line terminates
- traced: Heat/steam tracing if indicated

Return JSON: {"lines": [...], "count": N}
ONLY output the JSON, nothing else."""


COST_ESTIMATION_PROMPT = """You are an expert cost estimator for oil & gas and process engineering equipment. Based on the extracted P&ID data below, provide budget-level cost estimates for each item.

Use your knowledge of typical market prices for industrial equipment. Consider:
- Equipment type and typical sizing
- Material of construction (assume carbon steel unless specified otherwise)
- Standard industry pricing for oil & gas applications
- Prices should be in USD

EXTRACTED P&ID DATA:
{extracted_data}

Provide cost estimates in this exact JSON format:
{{
  "equipment_costs": [
    {{"tag": "...", "type": "...", "description": "...", "estimated_cost_usd": 0, "cost_basis": "brief explanation"}}
  ],
  "valve_costs": [
    {{"tag": "...", "type": "...", "size": "...", "estimated_cost_usd": 0, "cost_basis": "brief explanation"}}
  ],
  "instrument_costs": [
    {{"tag": "...", "type": "...", "function": "...", "estimated_cost_usd": 0, "cost_basis": "brief explanation"}}
  ],
  "piping_costs": [
    {{"line_number": "...", "size": "...", "estimated_cost_per_meter_usd": 0, "cost_basis": "brief explanation"}}
  ],
  "summary": {{
    "total_equipment": 0,
    "total_valves": 0,
    "total_instruments": 0,
    "total_piping_per_100m": 0,
    "grand_total_estimate": 0,
    "confidence_level": "low/medium/high",
    "assumptions": ["assumption 1", "assumption 2"],
    "exclusions": ["exclusion 1", "exclusion 2"]
  }}
}}

IMPORTANT:
- These are BUDGET estimates only (±30-50% accuracy)
- Exclude installation, engineering, freight costs
- Base on typical 2024-2026 market pricing
- If you cannot estimate, use 0 and explain in cost_basis
- Return ONLY the JSON, no other text."""


# ============================================================================
# IMAGE PROCESSING UTILITIES
# ============================================================================

def pdf_to_images(pdf_bytes: bytes, dpi: int = 200) -> List[Image.Image]:
    """Convert PDF pages to PIL Images"""
    if not PDF_SUPPORT:
        raise HTTPException(400, "PDF support not available. Install pdf2image and poppler.")
    
    try:
        images = convert_from_bytes(pdf_bytes, dpi=dpi)
        print(f"[PDF] Converted {len(images)} pages at {dpi} DPI")
        return images
    except Exception as e:
        raise HTTPException(500, f"PDF conversion failed: {str(e)}")


def image_to_base64(img: Image.Image, format: str = "PNG") -> str:
    """Convert PIL Image to base64 string"""
    buffer = io.BytesIO()
    # Convert to RGB if necessary (for RGBA images)
    if img.mode in ('RGBA', 'LA', 'P'):
        img = img.convert('RGB')
        format = "JPEG"
    img.save(buffer, format=format, quality=95)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')


def split_image_into_tiles(img: Image.Image, tiles: int = 4) -> List[Tuple[Image.Image, str]]:
    """
    Split image into tiles for detailed analysis.
    Returns list of (tile_image, position_label) tuples.
    
    For tiles=4: splits into 2x2 grid (top-left, top-right, bottom-left, bottom-right)
    For tiles=9: splits into 3x3 grid
    """
    width, height = img.size
    
    if tiles == 4:
        grid = 2
    elif tiles == 9:
        grid = 3
    else:
        grid = 2
    
    tile_width = width // grid
    tile_height = height // grid
    
    tiles_list = []
    positions = {
        (0, 0): "top-left",
        (1, 0): "top-right" if grid == 2 else "top-center",
        (2, 0): "top-right",
        (0, 1): "bottom-left" if grid == 2 else "middle-left",
        (1, 1): "bottom-right" if grid == 2 else "center",
        (2, 1): "middle-right",
        (0, 2): "bottom-left",
        (1, 2): "bottom-center",
        (2, 2): "bottom-right",
    }
    
    for row in range(grid):
        for col in range(grid):
            left = col * tile_width
            top = row * tile_height
            right = left + tile_width + (width % grid if col == grid - 1 else 0)
            bottom = top + tile_height + (height % grid if row == grid - 1 else 0)
            
            tile = img.crop((left, top, right, bottom))
            pos_label = positions.get((col, row), f"section-{col}-{row}")
            tiles_list.append((tile, pos_label))
    
    print(f"[TILE] Split image ({width}x{height}) into {len(tiles_list)} tiles")
    return tiles_list


def enhance_image_for_ocr(img: Image.Image) -> Image.Image:
    """
    Enhance image for better text/symbol recognition.
    - Increase contrast
    - Sharpen
    - Upscale if too small
    """
    from PIL import ImageEnhance, ImageFilter
    
    # Upscale small images
    width, height = img.size
    min_dimension = 1500
    if width < min_dimension or height < min_dimension:
        scale = max(min_dimension / width, min_dimension / height)
        new_size = (int(width * scale), int(height * scale))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
        print(f"[ENHANCE] Upscaled to {new_size}")
    
    # Enhance contrast
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.3)
    
    # Sharpen
    img = img.filter(ImageFilter.SHARPEN)
    
    return img


def should_use_tiling(img: Image.Image) -> bool:
    """Determine if image should be analyzed in tiles based on size/complexity"""
    width, height = img.size
    # Use tiling for large images (likely detailed P&IDs)
    return width > 2000 or height > 2000


# ============================================================================
# AI ANALYSIS FUNCTIONS
# ============================================================================

def analyze_image_with_vision(image_base64: str, prompt: str, media_type: str = "image/png") -> dict:
    """Send image to OpenAI Vision API with given prompt"""
    
    image_url = f"data:{media_type};base64,{image_base64}"
    
    response = get_openai_client().chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": image_url,
                            "detail": "high"
                        }
                    },
                    {"type": "text", "text": prompt}
                ]
            }
        ],
        max_tokens=4096,
        temperature=0.1
    )
    
    content = response.choices[0].message.content
    if not content:
        raise ValueError("Empty response from OpenAI")
    
    return parse_json_response(content)


def parse_json_response(content: str) -> dict:
    """Clean and parse JSON from AI response"""
    # Remove markdown code blocks
    if "```" in content:
        parts = content.split("```")
        for part in parts:
            part = part.strip()
            if part.startswith("json"):
                part = part[4:].strip()
            if part.startswith("{"):
                content = part
                break
    
    content = content.strip()
    
    # Find JSON object
    start = content.find("{")
    end = content.rfind("}") + 1
    if start != -1 and end > start:
        content = content[start:end]
    
    return json.loads(content)


def analyze_single_image_comprehensive(img: Image.Image) -> dict:
    """
    Analyze a single image comprehensively.
    For smaller/simpler images, use single-pass.
    For larger images, use tiled + focused extraction.
    """
    img = enhance_image_for_ocr(img)
    
    if should_use_tiling(img):
        print("[ANALYSIS] Using tiled analysis for large image")
        return analyze_with_tiling(img)
    else:
        print("[ANALYSIS] Using single-pass analysis")
        return analyze_single_pass(img)


def analyze_single_pass(img: Image.Image) -> dict:
    """Single comprehensive pass for smaller images"""
    img_b64 = image_to_base64(img)
    return analyze_image_with_vision(img_b64, PID_EXTRACTION_PROMPT)


def analyze_with_tiling(img: Image.Image) -> dict:
    """
    Analyze large image using:
    1. Full image overview pass (gets summary, notes, and initial extraction)
    2. Tiled detailed extraction for each category
    3. Merge and deduplicate results
    
    Falls back to single-pass if tiling fails completely.
    """
    results = {
        "equipment": [],
        "valves": [],
        "instruments": [],
        "lines": [],
        "notes": [],
        "summary": ""
    }
    
    success_count = 0
    
    # First: Overview pass on full image (resized for context)
    # This gets the summary AND initial extraction as fallback
    overview_img = img.copy()
    overview_img.thumbnail((2000, 2000), Image.Resampling.LANCZOS)
    overview_b64 = image_to_base64(overview_img)
    
    print("[ANALYSIS] Running overview pass...")
    overview_data = None
    try:
        overview_data = analyze_image_with_vision(overview_b64, PID_EXTRACTION_PROMPT)
        results["summary"] = overview_data.get("summary", "")
        results["notes"] = overview_data.get("notes", [])
        success_count += 1
        print(f"[ANALYSIS] Overview pass succeeded")
    except Exception as e:
        print(f"[ANALYSIS] Overview pass failed: {e}")
    
    # Second: Focused extraction passes on tiles
    tiles = split_image_into_tiles(img, tiles=4)
    
    focused_prompts = [
        ("equipment", PID_FOCUSED_EQUIPMENT_PROMPT),
        ("valves", PID_FOCUSED_VALVES_PROMPT),
        ("instruments", PID_FOCUSED_INSTRUMENTS_PROMPT),
        ("lines", PID_FOCUSED_LINES_PROMPT),
    ]
    
    tile_success = 0
    for tile_img, position in tiles:
        tile_b64 = image_to_base64(tile_img)
        
        for category, prompt in focused_prompts:
            print(f"[ANALYSIS] Extracting {category} from {position}...")
            try:
                tile_prompt = f"Analyzing the {position} section of a P&ID.\n\n{prompt}"
                tile_result = analyze_image_with_vision(tile_b64, tile_prompt)
                
                items = tile_result.get(category, [])
                for item in items:
                    item["_source_tile"] = position  # Track source for debugging
                results[category].extend(items)
                tile_success += 1
                
            except Exception as e:
                print(f"[ANALYSIS] {category} extraction from {position} failed: {e}")
    
    # If tiling completely failed but overview worked, use overview data
    if tile_success == 0 and overview_data:
        print("[ANALYSIS] Tiling failed, using overview extraction as fallback")
        results["equipment"] = overview_data.get("equipment", [])
        results["valves"] = overview_data.get("valves", [])
        results["instruments"] = overview_data.get("instruments", [])
        results["lines"] = overview_data.get("lines", [])
    elif tile_success == 0 and not overview_data:
        # Complete failure - try one more single pass
        print("[ANALYSIS] Complete failure, attempting final single-pass...")
        try:
            return analyze_single_pass(img)
        except Exception as e:
            print(f"[ANALYSIS] Final single-pass also failed: {e}")
            raise HTTPException(500, f"All analysis attempts failed: {e}")
    
    # Deduplicate results
    results["equipment"] = deduplicate_items(results["equipment"], key_field="tag")
    results["valves"] = deduplicate_items(results["valves"], key_field="tag")
    results["instruments"] = deduplicate_items(results["instruments"], key_field="tag")
    results["lines"] = deduplicate_items(results["lines"], key_field="line_number")
    
    # Clean up internal tracking fields
    for category in ["equipment", "valves", "instruments", "lines"]:
        for item in results[category]:
            item.pop("_source_tile", None)
    
    print(f"[ANALYSIS] Tiling complete: {tile_success} successful extractions")
    return results


def deduplicate_items(items: List[dict], key_field: str) -> List[dict]:
    """
    Remove duplicate items based on key field.
    Keeps the item with the most complete data.
    """
    if not items:
        return []
    
    seen = {}
    for item in items:
        key = item.get(key_field, "").strip().upper()
        if not key or key == "N/A":
            continue
        
        if key not in seen:
            seen[key] = item
        else:
            # Keep item with more non-empty fields
            existing_score = sum(1 for v in seen[key].values() if v and v != "N/A")
            new_score = sum(1 for v in item.values() if v and v != "N/A")
            if new_score > existing_score:
                seen[key] = item
    
    result = list(seen.values())
    print(f"[DEDUP] {len(items)} items -> {len(result)} unique")
    return result


def merge_extraction_results(results_list: List[dict]) -> dict:
    """Merge multiple extraction results (e.g., from multiple PDF pages)"""
    merged = {
        "equipment": [],
        "valves": [],
        "instruments": [],
        "lines": [],
        "notes": [],
        "summary": ""
    }
    
    summaries = []
    
    for result in results_list:
        merged["equipment"].extend(result.get("equipment", []))
        merged["valves"].extend(result.get("valves", []))
        merged["instruments"].extend(result.get("instruments", []))
        merged["lines"].extend(result.get("lines", []))
        merged["notes"].extend(result.get("notes", []))
        if result.get("summary"):
            summaries.append(result["summary"])
    
    # Deduplicate across all pages
    merged["equipment"] = deduplicate_items(merged["equipment"], "tag")
    merged["valves"] = deduplicate_items(merged["valves"], "tag")
    merged["instruments"] = deduplicate_items(merged["instruments"], "tag")
    merged["lines"] = deduplicate_items(merged["lines"], "line_number")
    merged["notes"] = list(set(merged["notes"]))  # Simple dedup for notes
    
    # Combine summaries
    if summaries:
        merged["summary"] = " | ".join(summaries)
    
    return merged


def estimate_costs(extracted_data: dict) -> dict:
    """Use AI to estimate costs based on extracted P&ID data"""
    
    prompt = COST_ESTIMATION_PROMPT.format(extracted_data=json.dumps(extracted_data, indent=2))
    
    response = get_openai_client().chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=4096,
        temperature=0.2
    )
    
    content = response.choices[0].message.content
    return parse_json_response(content)


# ============================================================================
# EXCEL REPORT GENERATION
# ============================================================================

def create_excel_report(data: dict, filename: str, cost_data: Optional[dict] = None) -> str:
    """Generate Excel report from extracted P&ID data with optional cost estimates"""
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    cost_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1, fill=header_fill):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
    
    def style_data(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left')
    
    # === EQUIPMENT Sheet ===
    ws_equip = wb.active
    ws_equip.title = "Equipment"
    headers = ["Tag", "Type", "Description", "Size/Capacity", "Material"]
    if cost_data:
        headers.extend(["Est. Cost (USD)", "Cost Basis"])
    ws_equip.append(headers)
    
    equipment_costs = {item.get("tag"): item for item in cost_data.get("equipment_costs", [])} if cost_data else {}
    
    for item in data.get("equipment", []):
        tag = item.get("tag", "")
        row = [tag, item.get("type", ""), item.get("description", ""), item.get("size", ""), item.get("material", "")]
        if cost_data:
            cost_item = equipment_costs.get(tag, {})
            row.extend([cost_item.get("estimated_cost_usd", 0), cost_item.get("cost_basis", "")])
        ws_equip.append(row)
    
    style_header(ws_equip)
    style_data(ws_equip)
    for i, width in enumerate([12, 18, 30, 18, 10, 15, 35] if cost_data else [12, 18, 30, 18, 10]):
        ws_equip.column_dimensions[get_column_letter(i+1)].width = width
    
    # === VALVES Sheet ===
    ws_valves = wb.create_sheet("Valves")
    headers = ["Tag", "Type", "Size", "Line Number", "Actuator", "Fail Position"]
    if cost_data:
        headers.extend(["Est. Cost (USD)", "Cost Basis"])
    ws_valves.append(headers)
    
    valve_costs = {item.get("tag"): item for item in cost_data.get("valve_costs", [])} if cost_data else {}
    
    for item in data.get("valves", []):
        tag = item.get("tag", "")
        row = [tag, item.get("type", ""), item.get("size", ""), item.get("line_number", ""),
               item.get("actuator", ""), item.get("fail_position", "")]
        if cost_data:
            cost_item = valve_costs.get(tag, {})
            row.extend([cost_item.get("estimated_cost_usd", 0), cost_item.get("cost_basis", "")])
        ws_valves.append(row)
    
    style_header(ws_valves)
    style_data(ws_valves)
    for i, width in enumerate([12, 15, 8, 15, 12, 10, 15, 30] if cost_data else [12, 15, 8, 15, 12, 10]):
        ws_valves.column_dimensions[get_column_letter(i+1)].width = width
    
    # === INSTRUMENTS Sheet ===
    ws_inst = wb.create_sheet("Instruments")
    headers = ["Tag", "Type", "Function", "Location", "Setpoint"]
    if cost_data:
        headers.extend(["Est. Cost (USD)", "Cost Basis"])
    ws_inst.append(headers)
    
    instrument_costs = {item.get("tag"): item for item in cost_data.get("instrument_costs", [])} if cost_data else {}
    
    for item in data.get("instruments", []):
        tag = item.get("tag", "")
        row = [tag, item.get("type", ""), item.get("function", ""), item.get("location", ""), item.get("setpoint", "")]
        if cost_data:
            cost_item = instrument_costs.get(tag, {})
            row.extend([cost_item.get("estimated_cost_usd", 0), cost_item.get("cost_basis", "")])
        ws_inst.append(row)
    
    style_header(ws_inst)
    style_data(ws_inst)
    for i, width in enumerate([12, 15, 15, 10, 12, 15, 30] if cost_data else [12, 15, 15, 10, 12]):
        ws_inst.column_dimensions[get_column_letter(i+1)].width = width
    
    # === LINES Sheet ===
    ws_lines = wb.create_sheet("Lines")
    headers = ["Line Number", "Size", "Service", "Spec", "Insulation", "From", "To"]
    if cost_data:
        headers.extend(["Est. Cost/m (USD)", "Cost Basis"])
    ws_lines.append(headers)
    
    piping_costs = {item.get("line_number"): item for item in cost_data.get("piping_costs", [])} if cost_data else {}
    
    for item in data.get("lines", []):
        line_num = item.get("line_number", "")
        row = [line_num, item.get("size", ""), item.get("service", ""), item.get("spec", ""),
               item.get("insulation", ""), item.get("from_equipment", item.get("from", "")), 
               item.get("to_equipment", item.get("to", ""))]
        if cost_data:
            cost_item = piping_costs.get(line_num, {})
            row.extend([cost_item.get("estimated_cost_per_meter_usd", 0), cost_item.get("cost_basis", "")])
        ws_lines.append(row)
    
    style_header(ws_lines)
    style_data(ws_lines)
    for i, width in enumerate([18, 8, 15, 10, 12, 15, 15, 12, 30] if cost_data else [18, 8, 15, 10, 12, 15, 15]):
        ws_lines.column_dimensions[get_column_letter(i+1)].width = width
    
    # === COST SUMMARY Sheet ===
    if cost_data:
        ws_cost = wb.create_sheet("Cost Summary")
        summary = cost_data.get("summary", {})
        
        ws_cost.append(["BUDGET COST ESTIMATE"])
        ws_cost['A1'].font = Font(bold=True, size=16, color="059669")
        ws_cost.append([])
        ws_cost.append(["Category", "Estimated Cost (USD)"])
        style_header(ws_cost, row=3, fill=cost_header_fill)
        
        ws_cost.append(["Equipment", summary.get("total_equipment", 0)])
        ws_cost.append(["Valves", summary.get("total_valves", 0)])
        ws_cost.append(["Instruments", summary.get("total_instruments", 0)])
        ws_cost.append(["Piping (per 100m)", summary.get("total_piping_per_100m", 0)])
        ws_cost.append([])
        ws_cost.append(["GRAND TOTAL", summary.get("grand_total_estimate", 0)])
        ws_cost['A9'].font = Font(bold=True, size=12)
        ws_cost['B9'].font = Font(bold=True, size=12, color="059669")
        
        ws_cost.append([])
        ws_cost.append(["Confidence Level:", summary.get("confidence_level", "N/A")])
        ws_cost.append([])
        ws_cost.append(["Key Assumptions:"])
        for assumption in summary.get("assumptions", []):
            ws_cost.append(["  •", assumption])
        ws_cost.append([])
        ws_cost.append(["Exclusions:"])
        for exclusion in summary.get("exclusions", []):
            ws_cost.append(["  •", exclusion])
        
        ws_cost.append([])
        ws_cost.append(["DISCLAIMER:"])
        ws_cost.append(["These are BUDGET estimates only with ±30-50% accuracy."])
        ws_cost.append(["For accurate pricing, obtain vendor quotes."])
        
        ws_cost.column_dimensions['A'].width = 25
        ws_cost.column_dimensions['B'].width = 50
        style_data(ws_cost, start_row=4)
    
    # === SUMMARY Sheet ===
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["P&ID Analysis Summary"])
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.append([])
    ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["Source File:", filename])
    ws_summary.append(["Parser Version:", "2.0 (Enhanced)"])
    ws_summary.append(["Cost Estimation:", "Included" if cost_data else "Not included"])
    ws_summary.append([])
    ws_summary.append(["Description:", data.get("summary", "N/A")])
    ws_summary.append([])
    ws_summary.append(["Item Counts:"])
    ws_summary.append(["  Equipment:", len(data.get("equipment", []))])
    ws_summary.append(["  Valves:", len(data.get("valves", []))])
    ws_summary.append(["  Instruments:", len(data.get("instruments", []))])
    ws_summary.append(["  Lines:", len(data.get("lines", []))])
    
    if cost_data:
        ws_summary.append([])
        total = cost_data.get("summary", {}).get("grand_total_estimate", 0)
        ws_summary.append(["Budget Estimate:", f"${total:,.0f}"])
    
    ws_summary.append([])
    ws_summary.append(["Notes:"])
    for note in data.get("notes", []):
        ws_summary.append(["  -", note])
    
    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 60
    
    # Save
    output_path = tempfile.mktemp(suffix=".xlsx")
    wb.save(output_path)
    return output_path


# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def home():
    """Serve the main page"""
    return FileResponse("static/index.html")


@app.post("/api/analyze")
async def analyze_pid(
    file: UploadFile = File(...),
    include_costs: Optional[str] = Form(default="false")
):
    """
    Analyze uploaded P&ID and return extracted data.
    
    Supports:
    - Images: JPG, PNG, WebP
    - PDFs: Single or multi-page (converted to images)
    
    Features:
    - Automatic tiling for large images
    - Multi-page PDF merging
    - Deduplication of extracted items
    - Optional cost estimation
    """
    
    # Check for demo mode
    demo_mode = os.environ.get("DEMO_MODE", "false").lower() == "true"
    
    # Validate file type
    content_type = file.content_type or ""
    allowed_types = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
    
    # Fallback detection by extension
    filename = file.filename or ""
    if not content_type or content_type == "application/octet-stream":
        ext = filename.lower().split(".")[-1] if "." in filename else ""
        content_type = {
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
            "webp": "image/webp",
            "pdf": "application/pdf"
        }.get(ext, content_type)
    
    if content_type not in allowed_types:
        raise HTTPException(400, f"Invalid file type '{content_type}'. Allowed: {', '.join(allowed_types)}")
    
    # Read file
    content = await file.read()
    
    # Check file size (max 50MB for PDFs, 20MB for images)
    max_size = 50 * 1024 * 1024 if "pdf" in content_type else 20 * 1024 * 1024
    if len(content) > max_size:
        raise HTTPException(400, f"File too large. Maximum size is {max_size // (1024*1024)}MB.")
    
    print(f"[UPLOAD] Received {filename} ({len(content)} bytes, {content_type})")
    
    try:
        if demo_mode:
            import time
            time.sleep(3)
            extracted_data = DEMO_MOCK_DATA
        else:
            # Process based on file type
            if "pdf" in content_type:
                # PDF: Convert to images and analyze each page
                if not PDF_SUPPORT:
                    raise HTTPException(400, 
                        "PDF support not available on this server. "
                        "Please convert your PDF to PNG/JPG and upload the image instead.")
                
                images = pdf_to_images(content, dpi=200)
                
                if len(images) == 1:
                    # Single page PDF
                    extracted_data = analyze_single_image_comprehensive(images[0])
                else:
                    # Multi-page PDF - analyze each and merge
                    print(f"[PDF] Processing {len(images)} pages...")
                    page_results = []
                    for i, img in enumerate(images):
                        print(f"[PDF] Analyzing page {i+1}/{len(images)}...")
                        page_result = analyze_single_image_comprehensive(img)
                        page_results.append(page_result)
                    
                    extracted_data = merge_extraction_results(page_results)
            else:
                # Image: Load and analyze
                img = Image.open(io.BytesIO(content))
                extracted_data = analyze_single_image_comprehensive(img)
        
        # Cost estimation if requested
        cost_data = None
        if include_costs.lower() == "true":
            print("[COST] Running cost estimation...")
            cost_data = estimate_costs(extracted_data)
        
        # Generate Excel report
        excel_path = create_excel_report(extracted_data, filename, cost_data)
        
        with open(excel_path, "rb") as f:
            excel_base64 = base64.b64encode(f.read()).decode('utf-8')
        
        os.unlink(excel_path)
        
        response_data = {
            "success": True,
            "data": extracted_data,
            "excel": excel_base64,
            "filename": f"PID_Extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "version": "2.0"
        }
        
        if cost_data:
            response_data["costs"] = cost_data
        
        # Stats for logging
        stats = {
            "equipment": len(extracted_data.get("equipment", [])),
            "valves": len(extracted_data.get("valves", [])),
            "instruments": len(extracted_data.get("instruments", [])),
            "lines": len(extracted_data.get("lines", []))
        }
        print(f"[RESULT] Extracted: {stats}")
        
        return response_data
        
    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON parse failed: {e}")
        raise HTTPException(500, f"Failed to parse AI response: {str(e)}")
    except Exception as e:
        print(f"[ERROR] Analysis failed: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Analysis failed: {str(e)}")


@app.get("/health")
async def health():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "P&ID Parser",
        "version": "2.0",
        "features": ["extraction", "cost_estimation", "pdf_support" if PDF_SUPPORT else "images_only", "tiling"]
    }


# Demo mock data
DEMO_MOCK_DATA = {
    "equipment": [
        {"tag": "V-101", "type": "Pressure Vessel", "description": "High Pressure Separator", "size": "2400mm x 6000mm", "material": "CS"},
        {"tag": "V-102", "type": "Pressure Vessel", "description": "Low Pressure Separator", "size": "1800mm x 4500mm", "material": "CS"},
        {"tag": "P-101A/B", "type": "Centrifugal Pump", "description": "Crude Oil Transfer Pump", "size": "150 m³/h", "material": "CS"},
        {"tag": "E-101", "type": "Heat Exchanger", "description": "Feed/Effluent Exchanger", "size": "500 m²", "material": "CS/SS"},
        {"tag": "C-101", "type": "Compressor", "description": "Gas Booster Compressor", "size": "2000 kW", "material": "CS"},
        {"tag": "T-101", "type": "Storage Tank", "description": "Crude Oil Storage Tank", "size": "5000 m³", "material": "CS"}
    ],
    "valves": [
        {"tag": "XV-101", "type": "Ball Valve", "size": "6\"", "line_number": "6\"-P-101", "actuator": "Pneumatic", "fail_position": "FC"},
        {"tag": "PV-102", "type": "Control Valve", "size": "4\"", "line_number": "4\"-P-102", "actuator": "Pneumatic", "fail_position": "FO"},
        {"tag": "PSV-101", "type": "Relief Valve", "size": "3\"", "line_number": "3\"-P-103", "actuator": "Self-acting", "fail_position": ""},
        {"tag": "CV-103", "type": "Check Valve", "size": "6\"", "line_number": "6\"-P-101", "actuator": "", "fail_position": ""},
        {"tag": "HV-104", "type": "Gate Valve", "size": "8\"", "line_number": "8\"-CW-201", "actuator": "Manual", "fail_position": ""},
        {"tag": "SDV-101", "type": "Shutdown Valve", "size": "6\"", "line_number": "6\"-P-101", "actuator": "Pneumatic", "fail_position": "FC"}
    ],
    "instruments": [
        {"tag": "PT-101", "type": "Pressure", "function": "Transmitter", "location": "Field", "setpoint": ""},
        {"tag": "LT-101", "type": "Level", "function": "Transmitter", "location": "Field", "setpoint": ""},
        {"tag": "FT-101", "type": "Flow", "function": "Transmitter", "location": "Field", "setpoint": ""},
        {"tag": "TT-101", "type": "Temperature", "function": "Transmitter", "location": "Field", "setpoint": ""},
        {"tag": "PI-102", "type": "Pressure", "function": "Indicator", "location": "Field", "setpoint": ""},
        {"tag": "LAH-101", "type": "Level", "function": "High Alarm", "location": "DCS", "setpoint": "80%"},
        {"tag": "PSHH-101", "type": "Pressure", "function": "High-High Switch", "location": "Field", "setpoint": "25 barg"}
    ],
    "lines": [
        {"line_number": "6\"-P-101", "size": "6\"", "service": "Crude Oil", "spec": "A1A", "insulation": "", "from_equipment": "V-101", "to_equipment": "P-101A/B"},
        {"line_number": "4\"-P-102", "size": "4\"", "service": "Produced Water", "spec": "A1A", "insulation": "", "from_equipment": "V-102", "to_equipment": "T-102"},
        {"line_number": "8\"-CW-201", "size": "8\"", "service": "Cooling Water", "spec": "B1A", "insulation": "", "from_equipment": "E-101", "to_equipment": "Utility Header"},
        {"line_number": "3\"-FG-101", "size": "3\"", "service": "Fuel Gas", "spec": "C1A", "insulation": "", "from_equipment": "V-101", "to_equipment": "Flare"},
        {"line_number": "10\"-P-103", "size": "10\"", "service": "Feed", "spec": "A1A", "insulation": "Hot", "from_equipment": "Header", "to_equipment": "V-101"}
    ],
    "notes": [
        "Design Pressure: 25 barg",
        "Design Temperature: 80°C",
        "Hazardous Area Classification: Zone 1"
    ],
    "summary": "This P&ID shows a typical oil & gas production facility with high and low pressure separators, transfer pumps, heat exchanger, and associated instrumentation. The system handles crude oil separation with produced water and gas handling facilities."
}


# ============================================================================
# RFQ AND DATASHEET ENDPOINTS
# ============================================================================

RFQ_EXTRACTION_PROMPT = """You are an expert tender/RFQ analyst for oil & gas, EPC, and industrial projects.

Analyze the uploaded RFQ/ITB/tender document and extract ALL relevant information.

Return a JSON object with these sections:
{
  "overview": {
    "document_type": "RFQ/ITB/Tender/Quote Request/etc.",
    "reference_number": "Document reference or tender number",
    "issuing_company": "Company issuing the RFQ",
    "project_name": "Project name if mentioned",
    "project_location": "Site/location"
  },
  "key_dates": {
    "submission_deadline": "Final submission date/time",
    "validity_period": "Quote validity required",
    "clarification_deadline": "Last date for questions",
    "delivery_date": "Required delivery date"
  },
  "scope_items": [
    {"item_no": "1", "description": "Item description", "quantity": "Qty with unit", "specifications": "Key specs"}
  ],
  "technical_requirements": {
    "standards": ["API 610", "ASME VIII", "etc."],
    "certifications": ["ISO 9001", "ATEX", "etc."],
    "documentation": ["GA drawings", "Test certs", "Manuals"]
  },
  "commercial_terms": {
    "payment_terms": "Payment conditions",
    "incoterms": "Delivery terms (FOB, CIF, etc.)",
    "warranty": "Warranty requirements"
  },
  "summary": "Brief summary of the RFQ"
}

Return ONLY the JSON."""


DATASHEET_EXTRACTION_PROMPT = """You are an expert equipment datasheet analyzer for industrial equipment.

Analyze the uploaded equipment datasheet and extract ALL specifications.

Return a JSON object:
{
  "general": {
    "manufacturer": "Equipment manufacturer/vendor",
    "model": "Model number",
    "tag_number": "Equipment tag if shown",
    "equipment_type": "Pump/Valve/Exchanger/Instrument/etc.",
    "service": "Service description"
  },
  "operating_conditions": {
    "design_pressure": "Design pressure with unit",
    "operating_pressure": "Operating pressure with unit",
    "design_temperature": "Design temperature with unit",
    "operating_temperature": "Operating temperature with unit",
    "flow_rate": "Flow rate with unit"
  },
  "physical": {
    "dimensions": "Overall dimensions",
    "weight": "Weight",
    "materials": "Key materials",
    "connections": "Inlet/outlet connections"
  },
  "all_specs": [
    {"parameter": "Parameter name", "value": "Value", "unit": "Unit"}
  ],
  "summary": "Brief summary"
}

Return ONLY the JSON."""


def create_rfq_excel(data: dict, filename: str) -> str:
    """Generate RFQ Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    overview = data.get("overview", {})
    ws.append(["RFQ Analysis"])
    ws['A1'].font = Font(bold=True, size=16)
    ws.append([])
    for k, v in overview.items():
        ws.append([k.replace("_", " ").title(), v])
    
    ws_s = wb.create_sheet("Scope Items")
    ws_s.append(["Item No", "Description", "Quantity", "Specifications"])
    for item in data.get("scope_items", []):
        ws_s.append([item.get("item_no", ""), item.get("description", ""),
                     item.get("quantity", ""), item.get("specifications", "")])
    
    output = tempfile.mktemp(suffix=".xlsx")
    wb.save(output)
    return output


def create_datasheet_excel(data: dict, filename: str) -> str:
    """Generate Datasheet Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"
    ws.append(["Datasheet Analysis"])
    ws['A1'].font = Font(bold=True, size=16)
    ws.append([])
    
    for section in ["general", "operating_conditions", "physical"]:
        ws.append([section.upper().replace("_", " ")])
        for k, v in data.get(section, {}).items():
            ws.append([k.replace("_", " ").title(), v])
        ws.append([])
    
    ws_a = wb.create_sheet("All Specs")
    ws_a.append(["Parameter", "Value", "Unit"])
    for spec in data.get("all_specs", []):
        ws_a.append([spec.get("parameter", ""), spec.get("value", ""), spec.get("unit", "")])
    
    output = tempfile.mktemp(suffix=".xlsx")
    wb.save(output)
    return output


@app.post("/api/parse-rfq")
async def parse_rfq(file: UploadFile = File(...)):
    """RFQ/Tender Analysis"""
    content_type = file.content_type or ""
    allowed = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
    if content_type not in allowed:
        raise HTTPException(400, f"Invalid file type")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "File too large")
    
    try:
        if "pdf" in content_type and PDF_SUPPORT:
            images = pdf_to_images(content, dpi=150)
            img_b64 = image_to_base64(images[0])
            media = "image/png"
        else:
            img_b64 = base64.b64encode(content).decode('utf-8')
            media = content_type
        
        extracted = analyze_image_with_vision(img_b64, RFQ_EXTRACTION_PROMPT, media)
        
        excel_path = create_rfq_excel(extracted, file.filename)
        with open(excel_path, "rb") as f:
            excel_b64 = base64.b64encode(f.read()).decode('utf-8')
        os.unlink(excel_path)
        
        return {
            "success": True,
            "data": extracted,
            "excel": excel_b64,
            "filename": f"RFQ_Extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    except Exception as e:
        raise HTTPException(500, f"Analysis failed: {e}")


@app.post("/api/parse-datasheet")
async def parse_datasheet(file: UploadFile = File(...)):
    """Equipment Datasheet Analysis"""
    content_type = file.content_type or ""
    allowed = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
    if content_type not in allowed:
        raise HTTPException(400, f"Invalid file type")
    
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "File too large")
    
    try:
        if "pdf" in content_type and PDF_SUPPORT:
            images = pdf_to_images(content, dpi=150)
            img_b64 = image_to_base64(images[0])
            media = "image/png"
        else:
            img_b64 = base64.b64encode(content).decode('utf-8')
            media = content_type
        
        extracted = analyze_image_with_vision(img_b64, DATASHEET_EXTRACTION_PROMPT, media)
        
        excel_path = create_datasheet_excel(extracted, file.filename)
        with open(excel_path, "rb") as f:
            excel_b64 = base64.b64encode(f.read()).decode('utf-8')
        os.unlink(excel_path)
        
        return {
            "success": True,
            "data": extracted,
            "excel": excel_b64,
            "filename": f"Datasheet_Extract_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    except Exception as e:
        raise HTTPException(500, f"Analysis failed: {e}")


# ============================================================================
# TENDER RESPONSE GENERATOR
# ============================================================================

TENDER_ANALYSIS_PROMPT = """You are an expert tender/RFQ analyst for oil & gas, EPC, and industrial projects.

Analyze this tender/RFQ document THOROUGHLY. Read ALL text including:
- Tables (line items, quantities, specifications)
- Requirements sections
- Technical specifications
- Commercial terms
- Deadlines and dates
- Compliance requirements

Extract EVERY detail into this JSON structure:

{
  "overview": {
    "document_type": "RFQ/ITB/Tender/etc.",
    "reference_number": "Document reference",
    "issuing_company": "Company name",
    "project_name": "Project name",
    "project_location": "Location",
    "issue_date": "When issued",
    "currency": "Currency for pricing"
  },
  "key_dates": {
    "submission_deadline": "Date and time",
    "clarification_deadline": "Last date for questions",
    "validity_period": "Quote validity required",
    "delivery_date": "Required delivery",
    "site_visit": "If applicable"
  },
  "scope_of_work": {
    "summary": "Brief description of overall scope",
    "line_items": [
      {
        "item_no": "1",
        "description": "Detailed item description",
        "quantity": "Qty with unit",
        "unit": "EA/SET/LOT/m/kg",
        "specifications": "Technical specs",
        "material": "If specified",
        "standard": "Applicable standard"
      }
    ],
    "services_included": ["Installation", "Commissioning", "Training", "etc."],
    "exclusions": ["What is NOT included"]
  },
  "technical_requirements": {
    "standards": ["API 610", "ASME B31.3", "etc."],
    "certifications": ["ISO 9001", "ATEX", "etc."],
    "testing": ["FAT", "Hydro test", "etc."],
    "documentation": ["Drawings", "MDR", "Test certs", "Manuals"],
    "materials": ["Specific material requirements"],
    "special_requirements": ["Any special technical needs"]
  },
  "commercial_terms": {
    "payment_terms": "Payment conditions",
    "incoterms": "FOB/CIF/DDP etc.",
    "warranty": "Warranty period required",
    "liquidated_damages": "LD clause details",
    "performance_bond": "If required",
    "insurance": "Insurance requirements",
    "price_validity": "How long prices must be valid"
  },
  "submission_requirements": {
    "format": "How to submit (email/portal/sealed)",
    "documents_required": ["Technical proposal", "Commercial proposal", "etc."],
    "contact_person": "Name and contact",
    "queries_to": "Who to send questions to"
  },
  "evaluation_criteria": [
    {"criterion": "Technical compliance", "weight": "40%"},
    {"criterion": "Price", "weight": "40%"}
  ],
  "compliance_checklist": [
    {"requirement": "Specific requirement from document", "mandatory": true}
  ],
  "risks_identified": [
    "Potential risk or concern noted in the tender"
  ],
  "summary": "Comprehensive 3-4 sentence summary of this tender"
}

Be EXHAUSTIVE - extract every line item, every specification, every date. Missing details = lost bid.
Return ONLY the JSON."""


TENDER_RESPONSE_PROMPT = """You are an expert bid manager for an engineering company (KU Automation) responding to tenders.

Based on the extracted tender requirements below, generate a PROFESSIONAL and DETAILED tender response.

EXTRACTED TENDER DATA:
{tender_data}

Generate a comprehensive response with these sections:

{{
  "executive_summary": {{
    "title": "Executive Summary",
    "content": "2-3 paragraphs introducing the company, expressing interest, and highlighting key strengths relevant to this tender. Be specific to the requirements."
  }},
  "technical_proposal": {{
    "title": "Technical Proposal",
    "scope_understanding": "Demonstrate understanding of the scope",
    "proposed_solution": "Detailed technical approach for each line item",
    "methodology": "How the work will be executed",
    "quality_assurance": "QA/QC approach",
    "schedule": "Proposed timeline with milestones"
  }},
  "compliance_matrix": {{
    "title": "Compliance Matrix",
    "items": [
      {{
        "requirement": "Requirement from tender",
        "compliance": "FULL/PARTIAL/DEVIATION",
        "response": "How we comply or explanation of deviation"
      }}
    ]
  }},
  "commercial_summary": {{
    "title": "Commercial Summary",
    "pricing_notes": "Notes about pricing approach",
    "payment_terms": "Proposed payment terms",
    "validity": "Price validity period",
    "delivery": "Delivery terms and timeline",
    "exclusions": ["What is not included in price"]
  }},
  "company_profile": {{
    "title": "Company Profile",
    "about": "Brief company introduction - KU Automation specializes in AI automation for engineering",
    "relevant_experience": ["Similar projects completed"],
    "certifications": ["ISO 9001", "etc."],
    "key_personnel": "Team that will execute"
  }},
  "risks_mitigations": {{
    "title": "Risk Assessment",
    "items": [
      {{
        "risk": "Identified risk",
        "mitigation": "How we will address it"
      }}
    ]
  }},
  "action_items": [
    "Clarifications needed from client",
    "Information to be provided"
  ]
}}

Make the response SPECIFIC to the tender requirements - not generic boilerplate.
Return ONLY the JSON."""


@app.post("/api/analyze-tender")
async def analyze_tender(file: UploadFile = File(...)):
    """Deep tender analysis - extracts all details from tender document"""
    content_type = file.content_type or ""
    filename = file.filename or ""
    
    # Fallback content type detection
    if not content_type or content_type == "application/octet-stream":
        ext = filename.lower().split(".")[-1] if "." in filename else ""
        content_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                        "webp": "image/webp", "pdf": "application/pdf"}.get(ext, content_type)
    
    allowed = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
    if content_type not in allowed:
        raise HTTPException(400, f"Invalid file type. Allowed: {', '.join(allowed)}")
    
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 50MB.")
    
    print(f"[TENDER] Analyzing: {filename} ({len(content)} bytes)")
    
    try:
        # Process file - handle multi-page PDFs
        if "pdf" in content_type and PDF_SUPPORT:
            images = pdf_to_images(content, dpi=200)
            print(f"[TENDER] PDF has {len(images)} pages")
            
            # For multi-page PDFs, analyze each page and merge
            all_extractions = []
            for i, img in enumerate(images[:10]):  # Max 10 pages
                print(f"[TENDER] Analyzing page {i+1}...")
                img_b64 = image_to_base64(img)
                try:
                    page_data = analyze_image_with_vision(img_b64, TENDER_ANALYSIS_PROMPT, "image/png")
                    all_extractions.append(page_data)
                except Exception as e:
                    print(f"[TENDER] Page {i+1} failed: {e}")
            
            # Merge all page extractions
            if all_extractions:
                extracted = merge_tender_extractions(all_extractions)
            else:
                raise HTTPException(500, "Failed to extract from any page")
        else:
            img_b64 = base64.b64encode(content).decode('utf-8')
            media = "application/pdf" if "pdf" in content_type else content_type
            extracted = analyze_image_with_vision(img_b64, TENDER_ANALYSIS_PROMPT, media)
        
        print(f"[TENDER] Extraction complete")
        
        return {
            "success": True,
            "data": extracted,
            "page_count": len(images) if "pdf" in content_type and PDF_SUPPORT else 1
        }
        
    except Exception as e:
        print(f"[TENDER ERROR] {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Analysis failed: {e}")


@app.post("/api/generate-tender-response")
async def generate_tender_response(
    file: UploadFile = File(None),
    tender_text: str = Form(None)
):
    """Generate AI tender response from document or text"""
    
    extracted_data = None
    
    # If file provided, analyze it first
    if file and file.filename:
        content_type = file.content_type or ""
        filename = file.filename or ""
        
        if not content_type or content_type == "application/octet-stream":
            ext = filename.lower().split(".")[-1] if "." in filename else ""
            content_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                            "webp": "image/webp", "pdf": "application/pdf"}.get(ext, content_type)
        
        content = await file.read()
        print(f"[TENDER-RESP] Processing file: {filename}")
        
        try:
            if "pdf" in content_type and PDF_SUPPORT:
                images = pdf_to_images(content, dpi=200)
                all_extractions = []
                for i, img in enumerate(images[:10]):
                    img_b64 = image_to_base64(img)
                    try:
                        page_data = analyze_image_with_vision(img_b64, TENDER_ANALYSIS_PROMPT, "image/png")
                        all_extractions.append(page_data)
                    except:
                        pass
                if all_extractions:
                    extracted_data = merge_tender_extractions(all_extractions)
            else:
                img_b64 = base64.b64encode(content).decode('utf-8')
                extracted_data = analyze_image_with_vision(img_b64, TENDER_ANALYSIS_PROMPT, content_type)
        except Exception as e:
            print(f"[TENDER-RESP] Extraction failed: {e}")
    
    # If text provided (or extraction failed), use that
    if not extracted_data and tender_text:
        # Try to parse as JSON first
        try:
            extracted_data = json.loads(tender_text)
        except:
            # Treat as raw text - create a simple structure
            extracted_data = {
                "overview": {"summary": "Tender requirements as provided"},
                "scope_of_work": {"summary": tender_text[:2000]},
                "raw_text": tender_text
            }
    
    if not extracted_data:
        raise HTTPException(400, "Please provide a file or tender text")
    
    print(f"[TENDER-RESP] Generating response...")
    
    # Generate the response
    try:
        prompt = TENDER_RESPONSE_PROMPT.format(tender_data=json.dumps(extracted_data, indent=2))
        
        response = get_openai_client().chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4096,
            temperature=0.3
        )
        
        response_data = parse_json_response(response.choices[0].message.content)
        
        print(f"[TENDER-RESP] Response generated")
        
        return {
            "success": True,
            "extracted": extracted_data,
            "response": response_data
        }
        
    except Exception as e:
        print(f"[TENDER-RESP ERROR] {e}")
        raise HTTPException(500, f"Response generation failed: {e}")


def merge_tender_extractions(extractions: List[dict]) -> dict:
    """Merge tender extractions from multiple pages"""
    if not extractions:
        return {}
    
    if len(extractions) == 1:
        return extractions[0]
    
    # Start with first page as base
    merged = extractions[0].copy()
    
    # Merge line items from all pages
    all_line_items = []
    all_compliance = []
    all_risks = []
    
    for ext in extractions:
        # Merge scope line items
        scope = ext.get("scope_of_work", {})
        if isinstance(scope, dict):
            items = scope.get("line_items", [])
            if isinstance(items, list):
                all_line_items.extend(items)
        
        # Merge compliance checklist
        compliance = ext.get("compliance_checklist", [])
        if isinstance(compliance, list):
            all_compliance.extend(compliance)
        
        # Merge risks
        risks = ext.get("risks_identified", [])
        if isinstance(risks, list):
            all_risks.extend(risks)
        
        # Update overview if empty in base
        overview = ext.get("overview", {})
        if isinstance(overview, dict):
            base_overview = merged.get("overview", {})
            for k, v in overview.items():
                if v and (not base_overview.get(k) or base_overview.get(k) == "N/A"):
                    base_overview[k] = v
            merged["overview"] = base_overview
        
        # Update key dates
        dates = ext.get("key_dates", {})
        if isinstance(dates, dict):
            base_dates = merged.get("key_dates", {})
            for k, v in dates.items():
                if v and (not base_dates.get(k) or base_dates.get(k) == "N/A"):
                    base_dates[k] = v
            merged["key_dates"] = base_dates
    
    # Deduplicate line items by item_no
    seen_items = {}
    for item in all_line_items:
        key = item.get("item_no", str(len(seen_items)))
        if key not in seen_items:
            seen_items[key] = item
    
    # Update merged with combined data
    if "scope_of_work" not in merged:
        merged["scope_of_work"] = {}
    merged["scope_of_work"]["line_items"] = list(seen_items.values())
    merged["compliance_checklist"] = list({str(c): c for c in all_compliance}.values())
    merged["risks_identified"] = list(set(all_risks))
    
    return merged


# ============================================================================
# PIPING DOCUMENT REVIEW
# ============================================================================

from piping_review import (
    PIPING_REVIEW_SYSTEM_PROMPT,
    get_review_prompt,
    get_document_types,
    PID_REVIEW_PROMPT,
    ISOMETRIC_REVIEW_PROMPT,
    PIPING_GA_REVIEW_PROMPT,
    PLOT_PLAN_REVIEW_PROMPT,
    LINE_LIST_REVIEW_PROMPT,
    STRESS_REPORT_REVIEW_PROMPT,
    GENERAL_PIPING_REVIEW_PROMPT
)


def create_review_excel(review_data: dict, filename: str) -> str:
    """Generate Excel comment register from review data"""
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hold_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    comment_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    note_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    
    # === COMMENTS Sheet ===
    ws = wb.active
    ws.title = "Review Comments"
    
    # Headers
    headers = ["No.", "Location", "Comment", "Code Reference", "Severity", "Category", "Contractor Response", "Status"]
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Comments
    comments = review_data.get("comments", [])
    for i, comment in enumerate(comments, 2):
        row_data = [
            comment.get("no", i-1),
            comment.get("location", ""),
            comment.get("comment", ""),
            comment.get("code_reference", ""),
            comment.get("severity", ""),
            comment.get("category", ""),
            "",  # Contractor Response (blank for them to fill)
            "OPEN"  # Status
        ]
        ws.append(row_data)
        
        # Apply severity highlighting
        severity = comment.get("severity", "").upper()
        fill = None
        if "HOLD" in severity:
            fill = hold_fill
        elif "COMMENT" in severity:
            fill = comment_fill
        elif "NOTE" in severity:
            fill = note_fill
        
        for col in range(1, 9):
            cell = ws.cell(row=i, column=col)
            cell.border = border
            cell.alignment = wrap_align
            if fill and col == 5:  # Severity column
                cell.fill = fill
    
    # Column widths
    widths = [6, 20, 50, 25, 12, 15, 30, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # === SUMMARY Sheet ===
    ws_sum = wb.create_sheet("Summary")
    doc_info = review_data.get("document_info", {})
    
    ws_sum.append(["DOCUMENT REVIEW SUMMARY"])
    ws_sum['A1'].font = Font(bold=True, size=16)
    ws_sum.append([])
    ws_sum.append(["Document Type:", doc_info.get("type", "N/A")])
    ws_sum.append(["Drawing Number:", doc_info.get("drawing_number", "N/A")])
    ws_sum.append(["Revision:", doc_info.get("revision", "N/A")])
    ws_sum.append(["Title:", doc_info.get("title", doc_info.get("line_number", "N/A"))])
    ws_sum.append([])
    ws_sum.append(["Review Date:", datetime.now().strftime("%Y-%m-%d")])
    ws_sum.append(["Reviewer:", "KU Automation AI Review"])
    ws_sum.append([])
    
    # Comment statistics
    hold_count = sum(1 for c in comments if "HOLD" in c.get("severity", "").upper())
    comment_count = sum(1 for c in comments if "COMMENT" in c.get("severity", "").upper())
    note_count = sum(1 for c in comments if "NOTE" in c.get("severity", "").upper())
    
    ws_sum.append(["COMMENT STATISTICS"])
    ws_sum['A11'].font = Font(bold=True)
    ws_sum.append(["Total Comments:", len(comments)])
    ws_sum.append(["🔴 HOLD:", hold_count])
    ws_sum.append(["🟡 COMMENT:", comment_count])
    ws_sum.append(["🟢 NOTE:", note_count])
    ws_sum.append([])
    
    ws_sum.append(["APPROVAL STATUS:", review_data.get("approval_status", "N/A")])
    ws_sum['B17'].font = Font(bold=True, size=12)
    
    # Readiness Score (new in v2.0)
    readiness = review_data.get("readiness_score", "N/A")
    ws_sum.append(["READINESS SCORE:", f"{readiness}%" if isinstance(readiness, (int, float)) else readiness])
    
    ws_sum.append([])
    ws_sum.append(["SUMMARY:"])
    ws_sum.append([review_data.get("summary", "N/A")])
    
    # Review Layers Assessment (new in v2.0)
    review_layers = review_data.get("review_layers", {})
    if review_layers:
        ws_sum.append([])
        ws_sum.append(["REVIEW LAYERS ASSESSMENT:"])
        layer_row = ws_sum.max_row
        ws_sum.cell(row=layer_row, column=1).font = Font(bold=True)
        for layer, status in review_layers.items():
            layer_name = layer.replace("_", " ").title()
            ws_sum.append([f"  {layer_name}:", status])
    
    # Critical Findings (new in v2.0)
    critical = review_data.get("critical_findings", [])
    if critical:
        ws_sum.append([])
        ws_sum.append(["CRITICAL FINDINGS (HOLD):"])
        cf_row = ws_sum.max_row
        ws_sum.cell(row=cf_row, column=1).font = Font(bold=True, color="CC0000")
        for finding in critical:
            ws_sum.append(["⚠️", finding])
    
    ws_sum.append([])
    ws_sum.append(["RECOMMENDATIONS:"])
    for rec in review_data.get("recommendations", []):
        ws_sum.append(["•", rec])
    
    ws_sum.column_dimensions['A'].width = 20
    ws_sum.column_dimensions['B'].width = 60
    
    # Save
    output = tempfile.mktemp(suffix=".xlsx")
    wb.save(output)
    return output


@app.post("/api/review-piping")
async def review_piping_document(
    file: UploadFile = File(...),
    doc_type: str = Form(default="auto"),
    project_context: str = Form(default="")
):
    """
    Review piping documents and generate engineering comments.
    
    doc_type options:
    - auto: Auto-detect document type
    - pid: P&ID
    - isometric / iso: Piping Isometric
    - ga / layout: Piping GA/Layout
    - line_list: Line List/Line Designation Table
    
    project_context: Optional project specifications to check against
    """
    
    content_type = file.content_type or ""
    filename = file.filename or ""
    
    # Fallback content type detection
    if not content_type or content_type == "application/octet-stream":
        ext = filename.lower().split(".")[-1] if "." in filename else ""
        content_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                        "webp": "image/webp", "pdf": "application/pdf"}.get(ext, content_type)
    
    allowed = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
    if content_type not in allowed:
        raise HTTPException(400, f"Invalid file type. Allowed: {', '.join(allowed)}")
    
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(400, "File too large. Max 50MB.")
    
    print(f"[PIPING REVIEW] Document: {filename}, Type: {doc_type}")
    
    try:
        # Get appropriate review prompt
        if doc_type.lower() == "auto":
            review_prompt = GENERAL_PIPING_REVIEW_PROMPT.format(project_context=project_context or "No specific context provided.")
        else:
            review_prompt = get_review_prompt(doc_type, project_context)
        
        # Process file
        all_comments = []
        combined_result = None
        
        if "pdf" in content_type and PDF_SUPPORT:
            images = pdf_to_images(content, dpi=200)
            print(f"[PIPING REVIEW] Processing {len(images)} pages")
            
            for i, img in enumerate(images[:10]):
                print(f"[PIPING REVIEW] Analyzing page {i+1}...")
                img_b64 = image_to_base64(img)
                
                try:
                    # Build messages with system prompt
                    messages = [
                        {"role": "system", "content": PIPING_REVIEW_SYSTEM_PROMPT},
                        {
                            "role": "user",
                            "content": [
                                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}", "detail": "high"}},
                                {"type": "text", "text": f"Page {i+1} of {len(images)}.\n\n{review_prompt}"}
                            ]
                        }
                    ]
                    
                    response = get_openai_client().chat.completions.create(
                        model="gpt-4o",
                        messages=messages,
                        max_tokens=4096,
                        temperature=0.2
                    )
                    
                    page_result = parse_json_response(response.choices[0].message.content)
                    
                    # Merge results
                    if combined_result is None:
                        combined_result = page_result
                    else:
                        # Add comments from this page
                        page_comments = page_result.get("comments", [])
                        for c in page_comments:
                            c["location"] = f"Page {i+1}: {c.get('location', '')}"
                        all_comments.extend(page_comments)
                    
                except Exception as e:
                    print(f"[PIPING REVIEW] Page {i+1} failed: {e}")
            
            # Merge all comments
            if combined_result and all_comments:
                combined_result["comments"] = combined_result.get("comments", []) + all_comments
                # Renumber comments
                for i, c in enumerate(combined_result["comments"], 1):
                    c["no"] = i
        
        else:
            # Single image
            img_b64 = base64.b64encode(content).decode('utf-8')
            media_type = content_type
            
            messages = [
                {"role": "system", "content": PIPING_REVIEW_SYSTEM_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{img_b64}", "detail": "high"}},
                        {"type": "text", "text": review_prompt}
                    ]
                }
            ]
            
            response = get_openai_client().chat.completions.create(
                model="gpt-4o",
                messages=messages,
                max_tokens=4096,
                temperature=0.2
            )
            
            combined_result = parse_json_response(response.choices[0].message.content)
        
        if not combined_result:
            raise HTTPException(500, "Failed to generate review")
        
        # Generate Excel
        excel_path = create_review_excel(combined_result, filename)
        with open(excel_path, "rb") as f:
            excel_b64 = base64.b64encode(f.read()).decode('utf-8')
        os.unlink(excel_path)
        
        # Stats
        comments = combined_result.get("comments", [])
        stats = {
            "total": len(comments),
            "hold": sum(1 for c in comments if "HOLD" in c.get("severity", "").upper()),
            "comment": sum(1 for c in comments if "COMMENT" in c.get("severity", "").upper()),
            "note": sum(1 for c in comments if "NOTE" in c.get("severity", "").upper())
        }
        
        print(f"[PIPING REVIEW] Complete: {stats}")
        
        return {
            "success": True,
            "data": combined_result,
            "statistics": stats,
            "excel": excel_b64,
            "filename": f"Review_Comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
        
    except Exception as e:
        print(f"[PIPING REVIEW ERROR] {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Review failed: {e}")


# ============================================================================
# TECHNICAL ASSISTANT — EXPENSES & INVOICE ENDPOINTS
# ============================================================================

import time
import requests as _requests
from io import BytesIO
from fastapi import Request as FARequest
from fastapi.responses import StreamingResponse

# ---------- FX Rate Cache (24h TTL) ----------
_fx_cache: Dict[str, Any] = {"rates": {}, "ts": 0}
_FX_TTL = 86400  # 24 hours

def _get_fx_rates() -> Dict[str, float]:
    """Return USD-base exchange rates dict; falls back to empty on error."""
    global _fx_cache
    now = time.time()
    if _fx_cache["ts"] and (now - _fx_cache["ts"]) < _FX_TTL and _fx_cache["rates"]:
        return _fx_cache["rates"]
    try:
        resp = _requests.get(
            "https://open.er-api.com/v6/latest/USD", timeout=5
        )
        data = resp.json()
        if data.get("result") == "success":
            _fx_cache["rates"] = data["rates"]
            _fx_cache["ts"] = now
            return _fx_cache["rates"]
    except Exception:
        pass
    # fallback: try exchangerate.host
    try:
        resp = _requests.get(
            "https://api.exchangerate.host/latest?base=USD", timeout=5
        )
        data = resp.json()
        if data.get("success"):
            _fx_cache["rates"] = data["rates"]
            _fx_cache["ts"] = now
            return _fx_cache["rates"]
    except Exception:
        pass
    return _fx_cache.get("rates", {})


def _usd_rate(currency: str, rates: Dict[str, float]) -> float:
    """Return units-per-USD for given ISO currency (i.e. how many X per 1 USD).
    To convert X -> USD: amount / rate_to_usd_denom
    Actually: rates are already relative to USD base.
    1 USD = rates[CUR]  =>  1 CUR = 1/rates[CUR] USD.
    We return the rate field as displayed: 'FX Rate' = rate of the currency vs USD
    matching the template (source_amount * fx_rate = usd_amount).
    So fx_rate = 1/rates[CUR] when base is USD.
    """
    cur = currency.upper().strip()
    if cur == "USD":
        return 1.0
    r = rates.get(cur)
    if r and r != 0:
        return round(1.0 / r, 6)
    return 1.0


# ---------- Expense extraction prompt ----------
EXPENSE_EXTRACTION_PROMPT = """You are an expert at extracting expense data from receipts, invoices, and expense documents.

Analyze the provided image/document and extract ALL expense items. Return ONLY valid JSON.

For each expense row return:
{
  "date": "YYYY-MM-DD or empty string if unclear",
  "description": "brief description of what was purchased",
  "currency": "ISO 4217 code (e.g. USD, GBP, EUR, AED, NGN)",
  "source_amount": numeric_float,
  "category": one of ["Flights & Hotels", "Transport", "Meals", "Other"],
  "vendor": "vendor/merchant name or null",
  "notes": "any extra notes or null"
}

Rules:
- Always output ISO 4217 currency codes (GBP not Pounds, EUR not Euro)
- If date is ambiguous, prefer DD/MM/YYYY interpretation for European/Middle East receipts
- Convert to YYYY-MM-DD format always
- Categorize intelligently: flights/hotels = Flights & Hotels, taxi/bus/train/car = Transport, food/restaurant = Meals, everything else = Other
- source_amount must be a number (no currency symbols)
- If a single receipt has multiple items, return them as separate rows

Return JSON: {"rows": [... list of expense objects ...]}
"""


# ---------- ReportLab PDF helpers ----------
def _make_expense_pdf(payload: dict) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    primary_blue = colors.HexColor('#1e3a5f')
    light_blue = colors.HexColor('#dbeafe')
    accent = colors.HexColor('#3b82f6')

    title_style = ParagraphStyle('Title', parent=styles['Normal'],
                                  fontSize=18, fontName='Helvetica-Bold',
                                  textColor=primary_blue, alignment=TA_CENTER,
                                  spaceAfter=8)
    label_style = ParagraphStyle('Label', parent=styles['Normal'],
                                  fontSize=8, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#6b7280'))
    value_style = ParagraphStyle('Value', parent=styles['Normal'],
                                  fontSize=9, fontName='Helvetica')
    small_bold = ParagraphStyle('SB', parent=styles['Normal'],
                                 fontSize=8, fontName='Helvetica-Bold')

    meta = payload.get('metadata', {})
    rows = payload.get('rows', [])
    name = meta.get('name', '')
    emp_id = meta.get('employee_id', '')
    period = meta.get('period', '')
    purpose = meta.get('purpose', '')
    date_prepared = meta.get('date_prepared', datetime.now().strftime('%d/%m/%Y'))

    # Categorise
    cats = {'Flights & Hotels': 0.0, 'Transport': 0.0, 'Meals': 0.0, 'Other': 0.0}
    for r in rows:
        cat = r.get('category', 'Other')
        cats[cat] = cats.get(cat, 0.0) + float(r.get('usd_amount', 0))
    grand_total = sum(cats.values())

    story = []
    story.append(Paragraph("Expense Summary", title_style))
    story.append(Spacer(1, 4*mm))

    # Header + Summary side by side
    header_data = [
        [Paragraph('<b>Date Prepared</b>', label_style), Paragraph(date_prepared, value_style),
         '', Paragraph('<b>Expenses at a glance</b>', small_bold), ''],
        [Paragraph('<b>Name</b>', label_style), Paragraph(name, value_style),
         '', Paragraph('Flights and Hotels', value_style), Paragraph(f"${cats['Flights & Hotels']:,.2f}", value_style)],
        [Paragraph('<b>Employee ID</b>', label_style), Paragraph(emp_id, value_style),
         '', Paragraph('Transport', value_style), Paragraph(f"${cats['Transport']:,.2f}", value_style)],
        [Paragraph('<b>Period Covered</b>', label_style), Paragraph(period, value_style),
         '', Paragraph('Meals', value_style), Paragraph(f"${cats['Meals']:,.2f}", value_style)],
        [Paragraph('<b>Purpose</b>', label_style), Paragraph(purpose, value_style),
         '', Paragraph('Other Expenses', value_style), Paragraph(f"${cats['Other']:,.2f}", value_style)],
        ['', '',
         '', Paragraph('<b>Total</b>', small_bold), Paragraph(f"<b>${grand_total:,.2f}</b>", small_bold)],
    ]
    header_table = Table(header_data, colWidths=[35*mm, 55*mm, 10*mm, 55*mm, 25*mm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (3, 0), (4, 0), light_blue),
        ('BACKGROUND', (3, 5), (4, 5), light_blue),
        ('FONTNAME', (3, 5), (4, 5), 'Helvetica-Bold'),
        ('GRID', (3, 0), (4, 5), 0.5, colors.HexColor('#93c5fd')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6*mm))

    # Main table
    col_headers = ['Date', 'Description', 'Curr', 'Source Amt', 'FX', 'Amount',
                   'Flights &\nHotels', 'Transport', 'Meals', 'Other', 'Total USD']
    table_data = [col_headers]

    cat_cols = {'Flights & Hotels': 6, 'Transport': 7, 'Meals': 8, 'Other': 9}

    for r in rows:
        cat = r.get('category', 'Other')
        usd = float(r.get('usd_amount', 0))
        row_data = [''] * 11
        row_data[0] = str(r.get('date', ''))
        row_data[1] = str(r.get('description', ''))
        row_data[2] = str(r.get('currency', 'USD'))
        row_data[3] = f"{float(r.get('source_amount', 0)):,.2f}"
        row_data[4] = f"{float(r.get('fx_rate', 1)):,.4f}"
        row_data[5] = f"{usd:,.2f}"
        col_idx = cat_cols.get(cat, 9)
        row_data[col_idx] = f"{usd:,.2f}"
        row_data[10] = f"{usd:,.2f}"
        table_data.append(row_data)

    # Totals row
    totals = ['', 'Total', '', '', '', '', 
              f"{cats['Flights & Hotels']:,.2f}",
              f"{cats['Transport']:,.2f}",
              f"{cats['Meals']:,.2f}",
              f"{cats['Other']:,.2f}",
              f"{grand_total:,.2f}"]
    table_data.append(totals)

    col_widths = [18*mm, 48*mm, 11*mm, 18*mm, 16*mm, 18*mm, 18*mm, 18*mm, 14*mm, 14*mm, 18*mm]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    main_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (2, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
        ('BACKGROUND', (0, -1), (-1, -1), light_blue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))
    story.append(main_table)
    story.append(Spacer(1, 8*mm))

    # Approver block
    approver_data = [
        [Paragraph('<b>Approved by:</b>', small_bold), '', ''],
        [Paragraph('Name:', label_style), Paragraph('_' * 30, value_style), ''],
        [Paragraph('Position:', label_style), Paragraph('_' * 30, value_style), ''],
        [Paragraph('Signature:', label_style), Paragraph('_' * 30, value_style), ''],
    ]
    appr_table = Table(approver_data, colWidths=[25*mm, 70*mm, 85*mm])
    appr_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(appr_table)

    doc.build(story)
    return buf.getvalue()


def _make_invoice_pdf(payload: dict) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    primary_blue = colors.HexColor('#1e3a5f')
    light_blue = colors.HexColor('#dbeafe')

    h1 = ParagraphStyle('H1', parent=styles['Normal'], fontSize=22, fontName='Helvetica-Bold',
                         textColor=primary_blue, alignment=TA_CENTER, spaceAfter=6)
    label_s = ParagraphStyle('LS', parent=styles['Normal'], fontSize=8,
                              fontName='Helvetica-Bold', textColor=colors.HexColor('#374151'))
    val_s = ParagraphStyle('VS', parent=styles['Normal'], fontSize=9, fontName='Helvetica')
    small_s = ParagraphStyle('SS', parent=styles['Normal'], fontSize=7, fontName='Helvetica',
                              textColor=colors.HexColor('#6b7280'))
    bold_s = ParagraphStyle('BS', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold')

    consultant = payload.get('consultant', {})
    invoice = payload.get('invoice', {})
    line_items = payload.get('line_items', [])
    reimbursables = payload.get('reimbursables', [])
    payment = payload.get('payment_terms', 'Payment to be made within 30 days of receipt of invoice')

    story = []
    story.append(Paragraph("INVOICE", h1))
    story.append(Spacer(1, 4*mm))

    # Consultant block top-left / Invoice no top-right
    consultant_text = f"""
    <b>{consultant.get('name', '')}</b><br/>
    {consultant.get('address', '').replace(chr(10), '<br/>')}<br/>
    <font color='#2563eb'>{consultant.get('email', '')}</font><br/>
    {consultant.get('phone', '')}
    """
    inv_no_block = [
        [Paragraph('<b>Invoice no.</b>', label_s), Paragraph(invoice.get('invoice_no', ''), bold_s)],
        [Paragraph('<b>Date</b>', label_s), Paragraph(invoice.get('date', ''), val_s)],
    ]
    inv_no_table = Table(inv_no_block, colWidths=[30*mm, 40*mm])
    inv_no_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (1, 0), (1, 0), light_blue),
    ]))

    top_data = [
        [Paragraph(consultant_text.strip(), val_s), inv_no_table],
    ]
    top_table = Table(top_data, colWidths=[110*mm, 75*mm])
    top_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(top_table)
    story.append(Spacer(1, 5*mm))

    # Invoice to + Reference
    client_addr = invoice.get('client_address', '').replace('\n', '<br/>')
    billing_data = [
        [Paragraph('<b>Invoice to</b>', label_s),
         Paragraph(f"{invoice.get('client_name', '')}<br/>{client_addr}", val_s)],
        [Paragraph('<b>Reference</b>', label_s),
         Paragraph(invoice.get('reference', ''), val_s)],
    ]
    bill_table = Table(billing_data, colWidths=[30*mm, 148*mm])
    bill_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
    ]))
    story.append(bill_table)
    story.append(Spacer(1, 5*mm))

    # Line items table
    li_headers = [Paragraph('<b>Description</b>', bold_s),
                  Paragraph('<b>Units</b>', bold_s),
                  Paragraph('<b>Daily Rate\nUSD</b>', bold_s),
                  Paragraph('<b>Total\nUSD</b>', bold_s)]
    li_data = [li_headers]
    for item in line_items:
        li_data.append([
            str(item.get('description', '')),
            str(item.get('units', '')),
            f"{float(item.get('daily_rate', 0)):,.2f}",
            f"{float(item.get('total', 0)):,.2f}",
        ])
    # Reimbursables header row if any
    if reimbursables:
        li_data.append([Paragraph('<b>Travel expenses</b>', bold_s), '', '', ''])
        for r in reimbursables:
            li_data.append([
                str(r.get('description', '')),
                str(r.get('units', '1.00')),
                f"{float(r.get('daily_rate', 0)):,.2f}",
                f"{float(r.get('total', 0)):,.2f}",
            ])

    grand_total = sum(float(i.get('total', 0)) for i in line_items) + \
                  sum(float(r.get('total', 0)) for r in reimbursables)
    li_data.append(['', '', Paragraph('<b>Total USD</b>', bold_s), Paragraph(f'<b>{grand_total:,.2f}</b>', bold_s)])

    li_table = Table(li_data, colWidths=[100*mm, 20*mm, 30*mm, 30*mm], repeatRows=1)
    li_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), primary_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
        ('BACKGROUND', (0, -1), (-1, -1), light_blue),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('SPAN', (0, -1), (1, -1)),
    ]))
    story.append(li_table)
    story.append(Spacer(1, 5*mm))

    # Signature
    story.append(Paragraph('<b>Signature:</b> ___________________________', label_s))
    story.append(Spacer(1, 6*mm))

    # Payment + Banking
    bank = consultant.get('bank', {})
    payment_text = f"""{payment}<br/><br/>Payment to be made by bank transfer to:"""
    bank_rows = [
        [Paragraph('<b>Payment</b>', label_s), Paragraph(payment_text, val_s)],
        [Paragraph('<b>Account Holder</b>', label_s), Paragraph(bank.get('account_holder', consultant.get('name', '')), val_s)],
        [Paragraph('<b>Name of Bank</b>', label_s), Paragraph(bank.get('bank_name', ''), val_s)],
        [Paragraph('<b>IBAN</b>', label_s), Paragraph(bank.get('iban', ''), val_s)],
        [Paragraph('<b>Account Number</b>', label_s), Paragraph(bank.get('account_number', ''), val_s)],
        [Paragraph('<b>SWIFT code</b>', label_s), Paragraph(bank.get('swift', ''), val_s)],
        [Paragraph('<b>Bank Address</b>', label_s), Paragraph(bank.get('bank_address', '').replace('\n', '<br/>'), val_s)],
    ]
    bank_table = Table(bank_rows, colWidths=[35*mm, 145*mm])
    bank_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    story.append(bank_table)

    doc.build(story)
    return buf.getvalue()


# ---------- Excel helpers ----------
def _make_expense_excel(payload: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    meta = payload.get('metadata', {})
    rows = payload.get('rows', [])
    name = meta.get('name', '')
    emp_id = meta.get('employee_id', '')
    period = meta.get('period', '')
    purpose = meta.get('purpose', '')
    date_prepared = meta.get('date_prepared', datetime.now().strftime('%d/%m/%Y'))

    # Styles
    blue_fill = PatternFill("solid", fgColor="1e3a5f")
    light_fill = PatternFill("solid", fgColor="dbeafe")
    grey_fill = PatternFill("solid", fgColor="f3f4f6")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="f9fafb")
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    label_font = Font(name='Calibri', bold=True, size=10)
    value_font = Font(name='Calibri', size=10)
    total_font = Font(name='Calibri', bold=True, size=10)
    thin = Side(style='thin', color='d1d5db')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'Expense Summary'
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='1e3a5f')
    ws['A1'].alignment = center
    ws['A1'].fill = light_fill
    ws.row_dimensions[1].height = 28

    # Header block (left side rows 3-7)
    header_labels = [
        ('A3', 'Date Prepared'), ('A4', 'Name'), ('A5', 'Employee ID'),
        ('A6', 'Period Covered'), ('A7', 'Purpose')
    ]
    header_values = [
        ('B3', date_prepared), ('B4', name), ('B5', emp_id),
        ('B6', period), ('B7', purpose)
    ]
    for cell, label in header_labels:
        ws[cell] = label
        ws[cell].font = label_font
        ws[cell].fill = grey_fill
        ws[cell].border = border
        ws[cell].alignment = left_align
    for cell, val in header_values:
        ws[cell] = val
        ws[cell].font = value_font
        ws[cell].border = border
        ws[cell].alignment = left_align
    ws.merge_cells('B3:D3')
    ws.merge_cells('B4:D4')
    ws.merge_cells('B5:D5')
    ws.merge_cells('B6:D6')
    ws.merge_cells('B7:D7')

    # Summary block (right side rows 3-8)
    cats = {'Flights & Hotels': 0.0, 'Transport': 0.0, 'Meals': 0.0, 'Other': 0.0}
    for r in rows:
        cat = r.get('category', 'Other')
        cats[cat] = cats.get(cat, 0.0) + float(r.get('usd_amount', 0))
    grand_total = sum(cats.values())

    ws.merge_cells('F3:K3')
    ws['F3'] = 'Expenses at a glance'
    ws['F3'].font = Font(name='Calibri', bold=True, size=11, color='1e3a5f')
    ws['F3'].fill = light_fill
    ws['F3'].alignment = center
    ws['F3'].border = border

    summary_rows = [
        ('F4', 'Flights and Hotels', 'K4', cats['Flights & Hotels']),
        ('F5', 'Transport', 'K5', cats['Transport']),
        ('F6', 'Meals', 'K6', cats['Meals']),
        ('F7', 'Other Expenses', 'K7', cats['Other']),
        ('F8', 'Total', 'K8', grand_total),
    ]
    for lc, label, vc, val in summary_rows:
        ws.merge_cells(f'{lc}:{chr(ord(lc[0])+4)}{lc[1:]}')
        ws[lc] = label
        ws[lc].font = label_font if lc != 'F8' else total_font
        ws[lc].fill = grey_fill if lc != 'F8' else light_fill
        ws[lc].border = border
        ws[lc].alignment = left_align
        ws[vc] = val
        ws[vc].font = value_font if lc != 'F8' else total_font
        ws[vc].number_format = '#,##0.00'
        ws[vc].border = border
        ws[vc].alignment = right_align
        ws[vc].fill = light_fill if lc == 'F8' else white_fill

    # Main table header (row 10)
    col_headers = ['Date', 'Description', 'Currency', 'Source Amount', 'FX', 'Amount (USD)',
                   'Flights & Hotels', 'Transport', 'Meals', 'Other', 'Total USD']
    for i, h in enumerate(col_headers, 1):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[10].height = 30

    # Column widths
    col_widths = [12, 35, 10, 15, 12, 14, 16, 12, 10, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    cat_cols = {'Flights & Hotels': 7, 'Transport': 8, 'Meals': 9, 'Other': 10}
    for ri, r in enumerate(rows):
        row_num = 11 + ri
        fill = white_fill if ri % 2 == 0 else alt_fill
        cat = r.get('category', 'Other')
        usd = float(r.get('usd_amount', 0))
        row_vals = [
            r.get('date', ''), r.get('description', ''), r.get('currency', 'USD'),
            float(r.get('source_amount', 0)), float(r.get('fx_rate', 1)), usd,
            None, None, None, None, usd
        ]
        cat_col_idx = cat_cols.get(cat, 10)
        row_vals[cat_col_idx - 1] = usd
        row_vals[10] = usd  # Total USD

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font = value_font
            cell.fill = fill
            cell.border = border
            cell.alignment = right_align if ci >= 4 else left_align
            if ci >= 4 and val is not None:
                cell.number_format = '#,##0.00'

    # Totals row
    total_row = 11 + len(rows)
    total_vals = ['', 'Total', '', '', '', '',
                  cats['Flights & Hotels'], cats['Transport'], cats['Meals'], cats['Other'], grand_total]
    for ci, val in enumerate(total_vals, 1):
        cell = ws.cell(row=total_row, column=ci, value=val)
        cell.font = total_font
        cell.fill = light_fill
        cell.border = border
        cell.alignment = right_align if ci >= 7 else left_align
        if ci >= 7 and isinstance(val, float):
            cell.number_format = '#,##0.00'

    # Approver block
    appr_row = total_row + 2
    ws.cell(row=appr_row, column=1, value='Approved by:').font = label_font
    for offset, label in enumerate([('Name', 1), ('Position', 2), ('Signature', 3)]):
        lbl, col_off = offset if isinstance(offset, tuple) else (offset, 0)
        r_num = appr_row + 1 + lbl
        pass
    for i, lbl in enumerate(['Name:', 'Position:', 'Signature:']):
        r_num = appr_row + 1 + i
        ws.cell(row=r_num, column=1, value=lbl).font = label_font
        ws.cell(row=r_num, column=2, value='').border = border
        ws.merge_cells(f'B{r_num}:E{r_num}')

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_invoice_excel(payload: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    consultant = payload.get('consultant', {})
    invoice = payload.get('invoice', {})
    line_items = payload.get('line_items', [])
    reimbursables = payload.get('reimbursables', [])
    payment = payload.get('payment_terms', 'Payment to be made within 30 days of receipt of invoice')

    blue_fill = PatternFill("solid", fgColor="1e3a5f")
    light_fill = PatternFill("solid", fgColor="dbeafe")
    grey_fill = PatternFill("solid", fgColor="f3f4f6")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="f9fafb")
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    label_font = Font(name='Calibri', bold=True, size=10)
    value_font = Font(name='Calibri', size=10)
    total_font = Font(name='Calibri', bold=True, size=11)
    thin = Side(style='thin', color='d1d5db')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'INVOICE'
    ws['A1'].font = Font(name='Calibri', bold=True, size=20, color='1e3a5f')
    ws['A1'].alignment = center
    ws['A1'].fill = light_fill
    ws.row_dimensions[1].height = 35

    # Consultant block (A3:C8)
    bank = consultant.get('bank', {})
    cons_fields = [
        consultant.get('name', ''),
        consultant.get('address', ''),
        consultant.get('email', ''),
        consultant.get('phone', ''),
    ]
    ws.merge_cells('A3:C3')
    ws['A3'] = 'Consultant'
    ws['A3'].font = label_font
    ws['A3'].fill = grey_fill
    ws['A3'].border = border
    for i, val in enumerate(cons_fields):
        r = 4 + i
        ws.merge_cells(f'A{r}:C{r}')
        ws[f'A{r}'] = val
        ws[f'A{r}'].font = value_font
        ws[f'A{r}'].border = border
        ws[f'A{r}'].alignment = left_align

    # Invoice No + Date block (E3:F4)
    ws['E3'] = 'Invoice no.'
    ws['F3'] = invoice.get('invoice_no', '')
    ws['E4'] = 'Date'
    ws['F4'] = invoice.get('date', '')
    for cell in ['E3', 'E4']:
        ws[cell].font = label_font
        ws[cell].fill = grey_fill
        ws[cell].border = border
    for cell in ['F3', 'F4']:
        ws[cell].font = value_font
        ws[cell].border = border
        ws[cell].alignment = right_align
    ws['F3'].fill = light_fill

    # Client address (A9:C11)
    ws['A9'] = 'Invoice to'
    ws['A9'].font = label_font
    ws['A9'].fill = grey_fill
    ws['A9'].border = border
    ws.merge_cells('B9:F11')
    ws['B9'] = f"{invoice.get('client_name', '')}\n{invoice.get('client_address', '')}"
    ws['B9'].font = value_font
    ws['B9'].border = border
    ws['B9'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells('A10:A11')
    ws.row_dimensions[9].height = 18
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 18

    # Reference (A12)
    ws['A12'] = 'Reference'
    ws['A12'].font = label_font
    ws['A12'].fill = grey_fill
    ws['A12'].border = border
    ws.merge_cells('B12:F12')
    ws['B12'] = invoice.get('reference', '')
    ws['B12'].font = value_font
    ws['B12'].border = border
    ws['B12'].alignment = left_align
    ws.row_dimensions[12].height = 18

    # Line items table (row 14+)
    li_header_row = 14
    li_cols = ['Description', 'Units', 'Daily Rate (USD)', 'Total (USD)']
    # Map to columns A, B, C(merged), D(merged), E, F
    ws.merge_cells(f'A{li_header_row}:C{li_header_row}')
    ws[f'A{li_header_row}'] = 'Description'
    ws[f'D{li_header_row}'] = 'Units'
    ws[f'E{li_header_row}'] = 'Daily Rate (USD)'
    ws[f'F{li_header_row}'] = 'Total (USD)'
    for col in ['A', 'D', 'E', 'F']:
        cell = ws[f'{col}{li_header_row}']
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = border

    current_row = li_header_row + 1
    for i, item in enumerate(line_items):
        fill = white_fill if i % 2 == 0 else alt_fill
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = str(item.get('description', ''))
        ws[f'A{current_row}'].font = value_font
        ws[f'A{current_row}'].border = border
        ws[f'A{current_row}'].fill = fill
        ws[f'A{current_row}'].alignment = left_align
        ws[f'D{current_row}'] = float(item.get('units', 0))
        ws[f'D{current_row}'].number_format = '0.00'
        ws[f'E{current_row}'] = float(item.get('daily_rate', 0))
        ws[f'E{current_row}'].number_format = '#,##0.00'
        ws[f'F{current_row}'] = float(item.get('total', 0))
        ws[f'F{current_row}'].number_format = '#,##0.00'
        for col in ['D', 'E', 'F']:
            ws[f'{col}{current_row}'].font = value_font
            ws[f'{col}{current_row}'].border = border
            ws[f'{col}{current_row}'].fill = fill
            ws[f'{col}{current_row}'].alignment = right_align
        current_row += 1

    if reimbursables:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = 'Travel expenses'
        ws[f'A{current_row}'].font = label_font
        ws[f'A{current_row}'].fill = grey_fill
        ws[f'A{current_row}'].border = border
        current_row += 1
        for i, r in enumerate(reimbursables):
            fill = white_fill if i % 2 == 0 else alt_fill
            ws.merge_cells(f'A{current_row}:C{current_row}')
            ws[f'A{current_row}'] = str(r.get('description', ''))
            ws[f'A{current_row}'].font = value_font
            ws[f'A{current_row}'].border = border
            ws[f'A{current_row}'].fill = fill
            ws[f'A{current_row}'].alignment = left_align
            ws[f'D{current_row}'] = float(r.get('units', 1))
            ws[f'D{current_row}'].number_format = '0.00'
            ws[f'E{current_row}'] = float(r.get('daily_rate', 0))
            ws[f'E{current_row}'].number_format = '#,##0.00'
            ws[f'F{current_row}'] = float(r.get('total', 0))
            ws[f'F{current_row}'].number_format = '#,##0.00'
            for col in ['D', 'E', 'F']:
                ws[f'{col}{current_row}'].font = value_font
                ws[f'{col}{current_row}'].border = border
                ws[f'{col}{current_row}'].fill = fill
                ws[f'{col}{current_row}'].alignment = right_align
            current_row += 1

    grand_total = sum(float(i.get('total', 0)) for i in line_items) + \
                  sum(float(r.get('total', 0)) for r in reimbursables)
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = 'Total USD'
    ws[f'A{current_row}'].font = total_font
    ws[f'A{current_row}'].fill = light_fill
    ws[f'A{current_row}'].border = border
    ws[f'A{current_row}'].alignment = right_align
    ws[f'F{current_row}'] = grand_total
    ws[f'F{current_row}'].font = total_font
    ws[f'F{current_row}'].number_format = '#,##0.00'
    ws[f'F{current_row}'].fill = light_fill
    ws[f'F{current_row}'].border = border
    ws[f'F{current_row}'].alignment = right_align

    # Signature block
    sig_row = current_row + 2
    ws[f'A{sig_row}'] = 'Signature:'
    ws[f'A{sig_row}'].font = label_font
    ws.merge_cells(f'B{sig_row}:D{sig_row}')
    ws[f'B{sig_row}'] = ''
    ws[f'B{sig_row}'].border = border

    # Payment + bank block
    pay_row = sig_row + 2
    payment_fields = [
        ('Payment Terms', payment),
        ('Account Holder', bank.get('account_holder', consultant.get('name', ''))),
        ('Name of Bank', bank.get('bank_name', '')),
        ('IBAN', bank.get('iban', '')),
        ('Account Number', bank.get('account_number', '')),
        ('SWIFT code', bank.get('swift', '')),
        ('Bank Address', bank.get('bank_address', '')),
    ]
    for i, (lbl, val) in enumerate(payment_fields):
        r = pay_row + i
        ws[f'A{r}'] = lbl
        ws[f'A{r}'].font = label_font
        ws[f'A{r}'].fill = grey_fill
        ws[f'A{r}'].border = border
        ws.merge_cells(f'B{r}:F{r}')
        ws[f'B{r}'] = val
        ws[f'B{r}'].font = value_font
        ws[f'B{r}'].border = border
        ws[f'B{r}'].alignment = left_align
        ws.row_dimensions[r].height = 20

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- API Endpoints ----------

@app.post("/api/expenses/extract")
async def expenses_extract(files: List[UploadFile] = File(...)):
    """
    Extract expense rows from one or more receipt files (PDF/JPG/PNG/WebP).
    Returns rows with FX rates pre-populated.
    """
    client = get_openai_client()
    rates = _get_fx_rates()
    all_rows = []

    for upload in files:
        content_type = upload.content_type or ""
        filename = upload.filename or ""
        if not content_type or content_type == "application/octet-stream":
            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
            content_type = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                            "png": "image/png", "webp": "image/webp",
                            "pdf": "application/pdf"}.get(ext, content_type)

        content = await upload.read()

        try:
            # Convert PDF first page to image if possible
            if "pdf" in content_type and PDF_SUPPORT:
                imgs = pdf_to_images(content, dpi=150)
                img_content = image_to_base64(imgs[0]) if imgs else None
                media_type = "image/png"
            elif "pdf" in content_type:
                # No pdf2image — try text extraction best-effort
                img_content = base64.b64encode(content).decode()
                media_type = "application/pdf"
            else:
                img_content = base64.b64encode(content).decode()
                media_type = content_type

            if not img_content:
                continue

            messages = [
                {
                    "role": "user",
                    "content": [
                        {"type": "image_url",
                         "image_url": {"url": f"data:{media_type};base64,{img_content}", "detail": "high"}},
                        {"type": "text", "text": EXPENSE_EXTRACTION_PROMPT}
                    ]
                }
            ]

            response = client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                max_tokens=2000,
                response_format={"type": "json_object"},
                temperature=0.1
            )

            result = json.loads(response.choices[0].message.content)
            file_rows = result.get("rows", [])

            # Enrich with FX rates
            for row in file_rows:
                cur = row.get("currency", "USD").upper()
                fx = _usd_rate(cur, rates)
                row["fx_rate"] = fx
                row["usd_amount"] = round(float(row.get("source_amount", 0)) * fx, 2)
                row["id"] = f"{time.time()}_{len(all_rows)}"

            all_rows.extend(file_rows)

        except Exception as e:
            print(f"[EXPENSE EXTRACT] Error on {filename}: {e}")
            import traceback; traceback.print_exc()

    return {"rows": all_rows, "fx_rates": {k: _usd_rate(k, rates) for k in rates}}


@app.post("/api/expenses/export-excel")
async def expenses_export_excel(request: FARequest):
    """
    Generate an .xlsx expense report matching the Kurt Bosse template.
    Body: { metadata: {...}, rows: [...] }
    """
    payload = await request.json()
    xlsx_bytes = _make_expense_excel(payload)
    meta = payload.get('metadata', {})
    name = meta.get('name', 'Report').replace(' ', '_')
    period = meta.get('period', datetime.now().strftime('%Y-%m')).replace('/', '-').replace(' ', '_')
    filename = f"Expenses_{name}_{period}.xlsx"
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/api/expenses/export-pdf")
async def expenses_export_pdf(request: FARequest):
    """
    Generate a PDF expense report matching the Kurt Bosse template.
    Body: { metadata: {...}, rows: [...] }
    """
    payload = await request.json()
    pdf_bytes = _make_expense_pdf(payload)
    meta = payload.get('metadata', {})
    name = meta.get('name', 'Report').replace(' ', '_')
    period = meta.get('period', datetime.now().strftime('%Y-%m')).replace('/', '-').replace(' ', '_')
    filename = f"Expenses_{name}_{period}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/api/invoice/export-excel")
async def invoice_export_excel(request: FARequest):
    """
    Generate an .xlsx invoice matching the Kurt Bosse invoice template.
    Body: { consultant: {...}, invoice: {...}, line_items: [...], reimbursables: [...], payment_terms: str }
    """
    payload = await request.json()
    xlsx_bytes = _make_invoice_excel(payload)
    inv = payload.get('invoice', {})
    inv_no = inv.get('invoice_no', 'INV').replace('/', '-')
    client = inv.get('client_name', 'Client').replace(' ', '_')
    filename = f"Invoice_{inv_no}_{client}.xlsx"
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/api/invoice/export-pdf")
async def invoice_export_pdf(request: FARequest):
    """
    Generate a PDF invoice matching the Kurt Bosse invoice template.
    Body: { consultant: {...}, invoice: {...}, line_items: [...], reimbursables: [...], payment_terms: str }
    """
    payload = await request.json()
    pdf_bytes = _make_invoice_pdf(payload)
    inv = payload.get('invoice', {})
    inv_no = inv.get('invoice_no', 'INV').replace('/', '-')
    client = inv.get('client_name', 'Client').replace(' ', '_')
    filename = f"Invoice_{inv_no}_{client}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
